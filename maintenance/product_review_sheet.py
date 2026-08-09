# READ-ONLY. Builds a fillable review workbook for the product attribute cleanup.
# Run inside odoo shell (the Odoo image ships openpyxl):
#   odoo shell -d tecas19 -c /etc/odoo/odoo.conf --no-http < product_review_sheet.py
# Output: /tmp/product_review.xlsx  -> copy out to backups/
#
# Fill the YOUR_DECISION / YOUR_VALUE columns and hand the file back; a later
# apply pass reads them. Nothing here writes to the database.
import sys, collections
sys.path.insert(0, "/tmp")
from product_lib import is_stuffed, classify, norm, FLAGGED_VALUE_HINTS

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = "/tmp/product_review.xlsx"

A = env["product.attribute"].with_context(active_test=False)
V = env["product.attribute.value"].with_context(active_test=False)
PTAV = env["product.template.attribute.value"].with_context(active_test=False)
T = env["product.template"].with_context(active_test=False)
SOL = env["sale.order.line"].with_context(active_test=False)

HEAD = Font(bold=True, color="FFFFFF")
HEAD_BG = PatternFill("solid", fgColor="0A2540")     # TECAS navy
ASK_BG = PatternFill("solid", fgColor="F5A623")      # amber = you fill this in
WARN_BG = PatternFill("solid", fgColor="FFE0E0")
WRAP = Alignment(vertical="top", wrap_text=True)

# ---------------------------------------------------------------- gather facts
val_templates = collections.defaultdict(set)
for p in PTAV.search([]):
    val_templates[p.product_attribute_value_id.id].add(p.product_tmpl_id.id)

# sale-line volume per template, so you can see what actually sells
sale_by_tmpl = collections.Counter()
for g in SOL.read_group([("product_id", "!=", False)], ["id"], ["product_id"]):
    pid = g["product_id"][0]
    tmpl = env["product.product"].with_context(active_test=False).browse(pid).product_tmpl_id
    if tmpl:
        sale_by_tmpl[tmpl.id] += g["product_id_count"]


def tmpl_brief(ids, limit=3):
    names, sales = [], 0
    for tid in sorted(ids):
        t = T.browse(tid)
        sales += sale_by_tmpl.get(tid, 0)
        if len(names) < limit:
            names.append(("%s%s" % (t.name or "?", "" if t.active else " [archived]"))[:60])
    extra = "" if len(ids) <= limit else "  (+%d more)" % (len(ids) - limit)
    return " | ".join(names) + extra, sales


def sheet(wb, title, headers, rows, ask_cols=(), widths=None, freeze="A2"):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font, cell.fill, cell.alignment = HEAD, HEAD_BG, WRAP
    for r in rows:
        ws.append(list(r))
    for c in ask_cols:
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=c).fill = ASK_BG
    for i, w in enumerate(widths or [], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    return ws


wb = Workbook()
wb.remove(wb.active)

# ------------------------------------------------------------ 1. REVIEW values
review_rows, auto_rows = [], []
for v in V.search([], order="attribute_id, name"):
    name = v.name or ""
    if not is_stuffed(name):
        continue
    tmpls = val_templates.get(v.id, set())
    if not tmpls:
        continue                      # unused -> handled by the T1 deletion tab
    prop, conf, why = classify(v.attribute_id.name, name)
    brief, sales = tmpl_brief(tmpls)
    row = [v.id, v.attribute_id.name, name, prop, why, len(tmpls), brief, sales,
           FLAGGED_VALUE_HINTS.get(prop, ""), "", ""]
    (auto_rows if conf == "AUTO" else review_rows).append(row)

HDR = ["value_id", "attribute", "current_value", "my_proposal", "why",
       "#templates", "templates_using_it", "sale_lines", "caution",
       "YOUR_DECISION", "YOUR_VALUE"]
W = [9, 18, 52, 22, 20, 10, 60, 10, 34, 16, 20]

ws = sheet(wb, "1. REVIEW (needs you)", HDR, review_rows, ask_cols=(10, 11), widths=W)
ws.insert_rows(1)
ws["A1"] = ("Stuffed attribute values my extractor could NOT resolve. "
            "Put keep / replace / delete in YOUR_DECISION; if replace, write the "
            "new value in YOUR_VALUE.")
ws["A1"].font = Font(bold=True)
ws.freeze_panes = "A3"

ws = sheet(wb, "2. AUTO (confirm)", HDR, auto_rows, ask_cols=(10, 11), widths=W)
ws.insert_rows(1)
ws["A1"] = ("Values I can resolve unambiguously. Default action is replace "
            "current_value with my_proposal. Leave YOUR_DECISION blank to accept; "
            "rows with a 'caution' note need a real look.")
ws["A1"].font = Font(bold=True)
ws.freeze_panes = "A3"
for r in range(3, ws.max_row + 1):
    if ws.cell(row=r, column=9).value:
        for c in range(1, len(HDR) + 1):
            ws.cell(row=r, column=c).fill = WARN_BG

# ------------------------------------------------------ 3. duplicate spellings
dup_rows = []
buckets = collections.defaultdict(list)
for v in V.search([]):
    buckets[(v.attribute_id.id, norm(v.name))].append(v)
for (aid, key), vs in buckets.items():
    if len(vs) < 2 or not key:
        continue
    used = [v for v in vs if val_templates.get(v.id)]
    sets = [val_templates.get(v.id, set()) for v in vs]
    overlap = set.intersection(*sets) if all(sets) else set()
    if len(used) <= 1:
        dec, why = "SAFE_DELETE_DUPE", "only one spelling is actually used"
    elif overlap:
        dec, why = "REVIEW", "both spellings on the SAME template — would collide"
    else:
        dec, why = "MERGE", "both in use on different templates — re-points variants"
    keeper = (used or vs)[0]
    for v in (used or vs):
        if (v.name or "").isupper() and not (keeper.name or "").isupper():
            continue
        keeper = v
        break
    for v in vs:
        if v.id == keeper.id:
            continue
        dup_rows.append([dec, A.browse(aid).name, keeper.id, keeper.name, v.id,
                         v.name, len(val_templates.get(v.id, ())), why, "", ""])

sheet(wb, "3. Duplicate spellings",
      ["decision", "attribute", "keep_id", "keep_value", "drop_id", "drop_value",
       "drop_used_by", "why", "YOUR_DECISION", "YOUR_VALUE"],
      sorted(dup_rows), ask_cols=(9, 10),
      widths=[18, 18, 9, 34, 9, 34, 12, 46, 16, 20])

# ------------------------------------------------------------ 4. T1 deletions
del_rows = []
for v in V.search([], order="attribute_id, name"):
    if not val_templates.get(v.id):
        del_rows.append(["value", v.id, v.attribute_id.name, v.name or "", "", ""])
for a in A.search([], order="name"):
    if not a.value_ids or all(not val_templates.get(v.id) for v in a.value_ids):
        del_rows.append(["attribute", a.id, a.name or "", "(%d values, all unused)"
                         % len(a.value_ids), "", ""])

ws = sheet(wb, "4. Unused - delete",
           ["kind", "id", "attribute", "value", "YOUR_DECISION", "notes"],
           del_rows, ask_cols=(5, 6), widths=[11, 9, 30, 56, 16, 40])
ws.insert_rows(1)
ws["A1"] = ("Referenced by NO product template. Deleting them changes nothing "
            "customers see. Blank YOUR_DECISION = delete; write keep to spare a row.")
ws["A1"].font = Font(bold=True)
ws.freeze_panes = "A3"

# ------------------------------------------------------------------ 0. summary
ws = wb.create_sheet("0. Read me", 0)
for line in [
    ["TECAS — product attribute cleanup, review sheet"],
    ["Generated read-only from prod (tecas19). Nothing has been changed yet."],
    [""],
    ["Tab", "What it is", "Rows", "What to do"],
    ["1. REVIEW (needs you)", "Stuffed values I could not resolve",
     len(review_rows), "Decide each: keep / replace / delete"],
    ["2. AUTO (confirm)", "Stuffed values with one clear token",
     len(auto_rows), "Skim; blank decision = accept my proposal"],
    ["3. Duplicate spellings", "Same value, different casing/spacing",
     len(dup_rows), "SAFE_DELETE rows are trivial; MERGE rows need a call"],
    ["4. Unused - delete", "Values/attributes no template references",
     len(del_rows), "Blank = delete; write keep to spare"],
    [""],
    ["Amber cells are the ones you fill in."],
    ["Duplicate product templates were surveyed and deliberately excluded: of 28"],
    ["duplicate-name groups, none had 2+ active templates, so there is nothing"],
    ["customer-visible to merge."],
]:
    ws.append(line)
ws["A1"].font = Font(bold=True, size=14)
for c in range(1, 5):
    ws.cell(row=4, column=c).font = HEAD
    ws.cell(row=4, column=c).fill = HEAD_BG
for i, w in enumerate([26, 44, 8, 52], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT)
print("=" * 70)
print("Review workbook written:", OUT)
print("  1. REVIEW (needs you)  :", len(review_rows), "rows")
print("  2. AUTO (confirm)      :", len(auto_rows), "rows")
print("  3. Duplicate spellings :", len(dup_rows), "rows")
print("  4. Unused (delete?)    :", len(del_rows), "rows")
print("=" * 70)
print("READ-ONLY — nothing was written to the database.")
